"""Импорт клиентской базы из Excel (.xlsx).

Поток: загрузка файла -> предпросмотр + автосопоставление колонок ->
подтверждение сопоставления и правил дублей -> импорт.
"""
from __future__ import annotations

import hashlib
from dataclasses import dataclass, field

from openpyxl import load_workbook

from ..models import Client, Funnel, ImportLog, Stage
from ..utils import normalize_phone, parse_ru_date, parse_ru_datetime

# Целевые поля CRM, на которые можно сопоставить колонку Excel
TARGET_FIELDS = [
    ("full_name", "Имя клиента (ФИО) / название сделки"),
    ("phone", "Телефон"),
    ("first_contact_date", "Дата обращения"),
    ("created_at", "Дата создания"),
    ("stage", "Стадия"),
    ("source", "База / источник"),
    ("funnel", "Воронка"),
    ("looking_for", "Что ищет / описание"),
    ("next_step", "Следующее действие"),
    ("task_due", "Дата задачи"),
    ("comment", "Комментарий"),
    ("manager", "Менеджер"),
    ("last_contact_at", "Дата последней коммуникации"),
    ("last_activity_at", "Последняя активность"),
]

# Ключевые слова в заголовке колонки -> целевое поле.
# Порядок важен: более специфичные подсказки проверяются раньше общих ("дата обращения" раньше "дата").
_HEADER_HINTS = {
    "full_name": ["фио", "имя клиент", "название сделки", "клиент", "имя", "name"],
    "phone": ["контакт", "телефон", "phone", "номер", "whatsapp"],
    "created_at": ["дата создания", "создан"],
    "last_contact_at": ["последней коммуникац", "коммуникац"],
    "last_activity_at": ["последняя активност", "активност"],
    "first_contact_date": ["дата обращен", "обращение"],
    "stage": ["стади", "статус", "этап", "stage"],
    "source": ["база", "источник", "source"],
    "funnel": ["воронка", "проект", "funnel"],
    "looking_for": ["что ищет", "что есть", "описание", "объект", "запрос"],
    "next_step": ["следующий шаг", "next", "действие"],
    "task_due": ["срок задачи", "срок", "задача", "дедлайн"],
    "comment": ["комментар", "заметк", "примечан", "comment"],
    "manager": ["менеджер", "ответствен", "manager"],
}

_STAGE_SYNONYMS = {
    "новая": "new", "новая заявка": "new", "заявка": "new",
    "принята": "accepted", "принят": "accepted", "принятый": "accepted",
    "консультация": "consult", "консультации": "consult",
    "горячий": "hot", "горячая": "hot",
    "холодный": "cold", "холодная": "cold",
    "заморожен": "frozen", "замороженный": "frozen", "заморозка": "frozen",
    "сделка": "won", "успех": "won", "успешно": "won",
    "проигранные": "lost", "проигранный": "lost", "проигран": "lost", "отказ": "lost",
}

_SOURCE_SYNONYMS = {
    "bitrix": "bitrix", "битрикс": "bitrix", "база bitrix": "bitrix", "база": "bitrix",
    "вручную": "manual", "manual": "manual",
    "excel": "excel", "импорт": "excel",
    "whatsapp": "whatsapp", "ватсап": "whatsapp",
    "звонок": "call", "call": "call",
    "instagram": "instagram", "инстаграм": "instagram",
    "рекомендация": "referral", "сарафан": "referral",
    "сайт": "site", "site": "site",
}


def _match_source(value) -> str:
    if not value:
        return Client.Source.EXCEL
    return _SOURCE_SYNONYMS.get(str(value).strip().lower(), Client.Source.EXCEL)


def _match_funnel(value, funnel_by_name: dict):
    if not value:
        return None
    return funnel_by_name.get(str(value).strip().lower())


# Выгрузки сделок из Bitrix24 часто кладут имя, источник и воронку в одну колонку:
# «Zaira Junusova - instagram Эл Насип» или «Гость - instagram Эл Насип».
_TITLE_SOURCE_WORDS = {
    "instagram": Client.Source.INSTAGRAM,
    "инстаграм": Client.Source.INSTAGRAM,
    "whatsapp": Client.Source.WHATSAPP,
    "ватсап": Client.Source.WHATSAPP,
    "звонок": Client.Source.CALL,
    "сайт": Client.Source.SITE,
    "bitrix": Client.Source.BITRIX,
    "битрикс": Client.Source.BITRIX,
}


def _parse_deal_title(title, funnels: list[Funnel]) -> tuple[str, str | None, Funnel | None]:
    """«Имя - instagram Эл Насип» -> («Имя», Source.INSTAGRAM, <Funnel Эл Насип>).

    Если ни источник, ни воронка внутри строки не распознаны (обычное «ФИО» без
    примесей) — возвращает исходную строку без изменений и None/None.
    """
    text = str(title).strip()
    low = text.lower()
    cut_at = len(text)
    found_source = None
    for word, src in _TITLE_SOURCE_WORDS.items():
        idx = low.find(word)
        if idx != -1:
            found_source = src
            cut_at = min(cut_at, idx)
    found_funnel = None
    for f in funnels:
        idx = low.find(f.name.lower())
        if idx != -1:
            found_funnel = f
            cut_at = min(cut_at, idx)
    if found_source is None and found_funnel is None:
        return text, None, None
    name = text[:cut_at].rstrip(" -–—").strip()
    return (name or text), found_source, found_funnel


@dataclass
class SheetPreview:
    sheet_name: str
    headers: list[str]
    header_row: int
    rows: list[list]           # первые строки данных для предпросмотра
    total_rows: int
    suggested_mapping: dict    # {target_field: column_index}


@dataclass
class ImportResult:
    total: int = 0
    created: int = 0
    updated: int = 0
    skipped: int = 0
    errors: list[str] = field(default_factory=list)
    # Пропущенные строки с причиной: {row, name, phone, reason}
    skipped_rows: list[dict] = field(default_factory=list)
    # Строки, отброшенные до обработки (пустые / без имени и телефона)
    empty_rows: list[int] = field(default_factory=list)
    # Клиенты, созданные без имени (телефон есть, ФИО пустое)
    noname_rows: list[dict] = field(default_factory=list)


def file_hash(file_bytes: bytes) -> str:
    return hashlib.sha256(file_bytes).hexdigest()


def _detect_header_row(ws, max_scan: int = 10):
    best_idx, best_score = 0, -1
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True)):
        filled = [c for c in row if c not in (None, "")]
        score = len(filled)
        # заголовок обычно текстовый
        text_cells = sum(1 for c in filled if isinstance(c, str))
        score += text_cells
        if score > best_score:
            best_score, best_idx = score, idx
    return best_idx  # 0-based


def _suggest_mapping(headers: list[str]) -> dict:
    mapping: dict[str, int] = {}
    used: set[int] = set()
    for target, hints in _HEADER_HINTS.items():
        for i, h in enumerate(headers):
            if i in used:
                continue
            hl = (h or "").strip().lower()
            if not hl:
                continue
            if any(hint in hl for hint in hints):
                mapping[target] = i
                used.add(i)
                break
    return mapping


def preview_workbook(file_bytes: bytes, preview_rows: int = 8) -> list[SheetPreview]:
    from io import BytesIO

    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    previews: list[SheetPreview] = []
    for ws in wb.worksheets:
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            continue
        hr = _detect_header_row(ws)
        headers_raw = all_rows[hr] if hr < len(all_rows) else ()
        headers = [str(c).strip() if c is not None else "" for c in headers_raw]
        # обрезаем хвост пустых колонок
        while headers and headers[-1] == "":
            headers.pop()
        data_rows = []
        total = 0
        for r in all_rows[hr + 1:]:
            cells = list(r[: len(headers)]) if headers else list(r)
            if not any(c not in (None, "") for c in cells):
                continue
            total += 1
            if len(data_rows) < preview_rows:
                data_rows.append([("" if c is None else c) for c in cells])
        previews.append(
            SheetPreview(
                sheet_name=ws.title,
                headers=headers,
                header_row=hr,
                rows=data_rows,
                total_rows=total,
                suggested_mapping=_suggest_mapping(headers),
            )
        )
    wb.close()
    return previews


def _match_stage(value, stage_cache: dict, default_stage: Stage) -> Stage:
    if not value:
        return default_stage
    key = str(value).strip().lower()
    slug = _STAGE_SYNONYMS.get(key)
    if slug and slug in stage_cache:
        return stage_cache[slug]
    # прямое совпадение по названию
    for st in stage_cache.values():
        if st.name.lower() == key:
            return st
    return default_stage


def _get_or_create_manager(name, manager_cache: dict):
    if not name:
        return None
    key = str(name).strip().lower()
    if not key:
        return None
    if key in manager_cache:
        return manager_cache[key]
    from django.contrib.auth import get_user_model

    User = get_user_model()
    display = str(name).strip()
    user = (
        User.objects.filter(first_name__iexact=display).first()
        or User.objects.filter(username__iexact=display).first()
    )
    if user is None:
        base = "".join(ch for ch in display.lower() if ch.isalnum()) or "manager"
        username = base
        n = 1
        while User.objects.filter(username=username).exists():
            n += 1
            username = f"{base}{n}"
        user = User.objects.create_user(username=username, first_name=display, role="manager")
        user.set_unusable_password()
        user.save()
    manager_cache[key] = user
    return user


def run_import(
    *,
    file_bytes: bytes,
    filename: str,
    sheet_name: str,
    header_row: int,
    mapping: dict,
    duplicate_strategy: str,   # skip | update | create
    default_manager=None,
    user=None,
    dry_run: bool = False,
) -> ImportResult:
    """mapping: {target_field: column_index}. Пустое значение — колонка не используется.

    dry_run=True — ничего не пишет в БД, только считает, что произойдёт.
    """
    from io import BytesIO

    result = ImportResult()
    default_stage = (
        Stage.objects.filter(is_active=True).order_by("order").first()
        or Stage.objects.order_by("order").first()
    )
    if default_stage is None:
        result.errors.append("В системе не настроены стадии — импорт невозможен.")
        return result

    stage_cache = {st.slug: st for st in Stage.objects.all()}
    manager_cache: dict = {}
    funnels = list(Funnel.objects.filter(is_active=True))
    funnel_by_name = {f.name.lower(): f for f in funnels}

    mapping = {k: int(v) for k, v in mapping.items() if v not in (None, "", "-")}

    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    def cell(row, field_name):
        idx = mapping.get(field_name)
        if idx is None or idx >= len(row):
            return None
        v = row[idx]
        return v

    seen_norms: dict[str, int] = {}  # норм. телефон -> номер строки в файле

    for offset, row in enumerate(rows[header_row + 1:]):
        excel_row = header_row + 2 + offset  # 1-based номер строки в Excel
        if not any(c not in (None, "") for c in row):
            result.empty_rows.append(excel_row)
            continue
        name = cell(row, "full_name")
        phone_raw = cell(row, "phone")
        name = str(name).strip() if name not in (None, "") else ""
        if not name and phone_raw in (None, ""):
            result.empty_rows.append(excel_row)
            continue
        result.total += 1

        # Bitrix и похожие выгрузки часто пишут «Имя - instagram Эл Насип» одной строкой —
        # вычленяем чистое имя, источник и воронку, если явные колонки для них не заданы.
        detected_source = None
        detected_funnel = None
        if name:
            cleaned_name, detected_source, detected_funnel = _parse_deal_title(name, funnels)
            if detected_source is not None or detected_funnel is not None:
                name = cleaned_name

        norm = normalize_phone(phone_raw)
        phone_str = str(phone_raw).strip() if phone_raw not in (None, "") else ""

        # дубль внутри самого файла
        if norm and norm in seen_norms:
            result.skipped += 1
            result.skipped_rows.append({
                "row": excel_row, "name": name or "(без имени)", "phone": phone_str,
                "reason": f"тот же телефон, что в строке {seen_norms[norm]} этого файла",
            })
            continue
        if norm:
            seen_norms[norm] = excel_row

        manager = _get_or_create_manager(cell(row, "manager"), manager_cache) or default_manager
        stage = _match_stage(cell(row, "stage"), stage_cache, default_stage)

        source_value = _match_source(cell(row, "source")) if mapping.get("source") else (
            detected_source or Client.Source.EXCEL
        )
        funnel_value = _match_funnel(cell(row, "funnel"), funnel_by_name) if mapping.get("funnel") else detected_funnel

        fields = dict(
            full_name=name or "Без имени",
            phone=phone_str,
            first_contact_date=parse_ru_date(cell(row, "first_contact_date")),
            looking_for=str(cell(row, "looking_for") or "").strip(),
            next_step=str(cell(row, "next_step") or "").strip()[:255],
            comment=str(cell(row, "comment") or "").strip(),
            source=source_value,
            funnel=funnel_value,
            last_contact_at=parse_ru_datetime(cell(row, "last_contact_at")),
            last_activity_at=parse_ru_datetime(cell(row, "last_activity_at")),
        )
        created_at_value = parse_ru_datetime(cell(row, "created_at"))

        existing = None
        if norm:
            existing = Client.objects.filter(phone_normalized=norm).first()
        elif name:
            # без телефона сверяем по имени
            existing = Client.objects.filter(full_name__iexact=name).first()

        if existing:
            if duplicate_strategy == "skip":
                result.skipped += 1
                result.skipped_rows.append({
                    "row": excel_row, "name": name or "(без имени)", "phone": phone_str,
                    "reason": (
                        f"клиент с таким телефоном уже есть в базе: "
                        f"«{existing.full_name}» (ID {existing.id})"
                        if norm else
                        f"клиент с таким именем уже есть в базе (ID {existing.id})"
                    ),
                })
                continue
            if duplicate_strategy == "update":
                if not dry_run:
                    for k, v in fields.items():
                        if v:
                            setattr(existing, k, v)
                    if manager and not existing.manager_id:
                        existing.manager = manager
                    existing.save()
                    _history(existing, "import", f"Данные обновлены при импорте ({filename})", user)
                result.updated += 1
                continue
            # create -> падаем ниже

        if not name:
            result.noname_rows.append({"row": excel_row, "phone": phone_str, "id": None})

        if not dry_run:
            client = Client.objects.create(
                stage=stage, manager=manager, created_by=user, **fields
            )
            if created_at_value:
                # created_at — auto_now_add, обычным save() не переопределяется.
                Client.objects.filter(pk=client.pk).update(created_at=created_at_value)
            _history(client, "import", f"Импортирован из {filename}", user)
            if not name and result.noname_rows:
                result.noname_rows[-1]["id"] = client.id

            task_due = parse_ru_date(cell(row, "task_due"))
            if task_due and client.next_step:
                from ..models import Task

                Task.objects.create(
                    title=client.next_step[:255],
                    client=client,
                    manager=manager,
                    due_date=task_due,
                    created_by=user,
                )
        result.created += 1

    if dry_run:
        return result

    ImportLog.objects.create(
        user=user,
        filename=filename,
        file_hash=file_hash(file_bytes),
        mapping={k: v for k, v in mapping.items()},
        total_rows=result.total,
        created_count=result.created,
        updated_count=result.updated,
        skipped_count=result.skipped,
        status=ImportLog.Status.DONE,
    )
    return result


def _history(client, kind, text, user):
    from ..models import ClientHistory

    ClientHistory.objects.create(client=client, kind=kind, text=text, user=user)
