"""Экспорт клиентов в Excel (.xlsx)."""
from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font

from ..utils import phone_display

COLUMNS = [
    ("ФИО", lambda c: c.full_name),
    ("Телефон", lambda c: phone_display(c.phone)),
    ("Доп. телефон", lambda c: c.phone_extra),
    ("WhatsApp", lambda c: phone_display(c.whatsapp_phone or c.phone)),
    ("Дата обращения", lambda c: c.first_contact_date.strftime("%d.%m.%Y") if c.first_contact_date else ""),
    ("Стадия", lambda c: c.stage.name if c.stage_id else ""),
    ("Что ищет", lambda c: c.looking_for),
    ("Что есть", lambda c: c.what_has),
    ("Бюджет", lambda c: c.budget),
    ("Следующий шаг", lambda c: c.next_step),
    ("Срок след. действия", lambda c: c.next_step_at.strftime("%d.%m.%Y %H:%M") if c.next_step_at else ""),
    ("Менеджер", lambda c: c.manager.display_name if c.manager_id else ""),
    ("Источник", lambda c: c.get_source_display()),
    ("Комментарий", lambda c: c.comment),
    ("Создан", lambda c: c.created_at.strftime("%d.%m.%Y")),
]


def export_clients(queryset) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Клиенты"

    ws.append([title for title, _ in COLUMNS])
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for client in queryset.select_related("stage", "manager"):
        ws.append([getter(client) for _, getter in COLUMNS])

    widths = [24, 18, 16, 18, 15, 14, 30, 30, 18, 24, 20, 18, 16, 40, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
